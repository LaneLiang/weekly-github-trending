import json
from collections import defaultdict
from pathlib import Path

from lanes_ceo.storage.sqlite_store import SQLiteStore


class ScoreReporter:
    """Exports score records to Excel and JSON from the SQLite runtime store.

    Fraction.xlsx is the human-readable report; score_records.json is the
    machine-readable source. Both are generated from the same SQLite data.
    """

    def __init__(self, store: SQLiteStore) -> None:
        self._store = store

    def export_json(self, output_dir: str | Path) -> Path:
        output = Path(output_dir) / "score_records.json"
        records = self._store.list_score_records()
        output.parent.mkdir(parents=True, exist_ok=True)
        output.write_text(
            json.dumps(
                [
                    {
                        "score_id": r.score_id,
                        "job_id": r.job_id,
                        "scored_role": r.scored_role,
                        "scorer_role": r.scorer_role,
                        "score": r.score,
                        "rating": r.rating,
                        "review_summary": r.review_summary,
                        "created_at": r.created_at,
                        "month_bucket": r.month_bucket,
                    }
                    for r in records
                ],
                indent=2,
                ensure_ascii=False,
            ),
            encoding="utf-8",
        )
        return output

    def export_xlsx(self, output_dir: str | Path) -> Path:
        """Export score records to Fraction.xlsx with Dashboard, Records, and Monthly sheets."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        records = self._store.list_score_records()
        output = Path(output_dir) / "Fraction.xlsx"
        output.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        center_align = Alignment(horizontal="center", vertical="center")

        def _style_header(ws, row, ncols):
            for col in range(1, ncols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align

        def _auto_width(ws, ncols, max_rows=None):
            for col in range(1, ncols + 1):
                max_len = 0
                for row in ws.iter_rows(min_col=col, max_col=col, values_only=True):
                    for val in row:
                        if val:
                            max_len = max(max_len, len(str(val)))
                ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 50)

        # ── Sheet 1: Dashboard ──
        ws1 = wb.active
        ws1.title = "Dashboard"
        scores = [r.score for r in records] if records else [0]
        ratings: dict[str, int] = defaultdict(int)
        months: dict[str, list[int]] = defaultdict(list)
        for r in records:
            ratings[r.rating] += 1
            months[r.month_bucket].append(r.score)

        ws1.merge_cells("A1:D1")
        ws1.cell(row=1, column=1, value="LANEs_CEO Score Dashboard").font = Font(bold=True, size=16)

        ws1.cell(row=3, column=1, value="Total Jobs").font = header_font
        ws1.cell(row=3, column=2, value=len(records))
        ws1.cell(row=4, column=1, value="Average Score").font = header_font
        ws1.cell(row=4, column=2, value=round(sum(scores) / len(scores), 1))
        ws1.cell(row=5, column=1, value="Pass Rate").font = header_font
        pass_count = sum(1 for r in records if r.score >= 80)
        ws1.cell(row=5, column=2, value=f"{pass_count}/{len(records)} ({round(pass_count/len(records)*100, 1) if records else 0}%)")

        # Rating distribution
        ws1.cell(row=7, column=1, value="Rating").font = header_font
        ws1.cell(row=7, column=2, value="Count").font = header_font
        for i, (rating, count) in enumerate(sorted(ratings.items()), 8):
            ws1.cell(row=i, column=1, value=rating)
            ws1.cell(row=i, column=2, value=count)

        # Monthly trend
        monthly_start = 7
        ws1.cell(row=monthly_start, column=4, value="Monthly Trend").font = header_font
        ws1.cell(row=monthly_start + 1, column=4, value="Month").font = header_font
        ws1.cell(row=monthly_start + 1, column=5, value="Jobs").font = header_font
        ws1.cell(row=monthly_start + 1, column=6, value="Avg Score").font = header_font
        ws1.cell(row=monthly_start + 1, column=7, value="Pass Rate").font = header_font
        row = monthly_start + 2
        for month, mscores in sorted(months.items()):
            mp = sum(1 for s in mscores if s >= 80)
            ws1.cell(row=row, column=4, value=month)
            ws1.cell(row=row, column=5, value=len(mscores))
            ws1.cell(row=row, column=6, value=round(sum(mscores) / len(mscores), 1))
            ws1.cell(row=row, column=7, value=f"{round(mp/len(mscores)*100, 1)}%")
            row += 1

        _auto_width(ws1, 7)

        # ── Sheet 2: Score Records ──
        ws2 = wb.create_sheet("Score Records")
        s2_headers = ["Score ID", "Job ID", "Role", "Reviewer", "Score", "Rating", "Month", "Created At", "Summary"]
        for c, h in enumerate(s2_headers, 1):
            ws2.cell(row=1, column=c, value=h)
        _style_header(ws2, 1, len(s2_headers))

        for i, r in enumerate(records, 2):
            ws2.cell(row=i, column=1, value=r.score_id)
            ws2.cell(row=i, column=2, value=r.job_id)
            ws2.cell(row=i, column=3, value=r.scored_role)
            ws2.cell(row=i, column=4, value=r.scorer_role)
            ws2.cell(row=i, column=5, value=r.score)
            ws2.cell(row=i, column=6, value=r.rating)
            ws2.cell(row=i, column=7, value=r.month_bucket)
            ws2.cell(row=i, column=8, value=r.created_at)
            ws2.cell(row=i, column=9, value=r.review_summary[:200] if r.review_summary else "")

        ws2.freeze_panes = "A2"
        _auto_width(ws2, len(s2_headers))

        # ── Sheet 3: Monthly ──
        ws3 = wb.create_sheet("Monthly")
        m3_headers = ["Month", "Total Jobs", "Avg Score", "Pass Count", "Fail Count", "Pass Rate"]
        for c, h in enumerate(m3_headers, 1):
            ws3.cell(row=1, column=c, value=h)
        _style_header(ws3, 1, len(m3_headers))

        for i, (month, mscores) in enumerate(sorted(months.items()), 2):
            mp = sum(1 for s in mscores if s >= 80)
            ws3.cell(row=i, column=1, value=month)
            ws3.cell(row=i, column=2, value=len(mscores))
            ws3.cell(row=i, column=3, value=round(sum(mscores) / len(mscores), 1))
            ws3.cell(row=i, column=4, value=mp)
            ws3.cell(row=i, column=5, value=len(mscores) - mp)
            ws3.cell(row=i, column=6, value=f"{round(mp/len(mscores)*100, 1) if mscores else 0}%")

        _auto_width(ws3, len(m3_headers))

        wb.save(str(output))
        return output

    def monthly_summary(self) -> dict:
        records = self._store.list_score_records()
        if not records:
            return {"total_jobs": 0, "avg_score": 0, "ratings": {}}
        scores = [r.score for r in records]
        ratings: dict[str, int] = {}
        for r in records:
            ratings[r.rating] = ratings.get(r.rating, 0) + 1
        return {
            "total_jobs": len(records),
            "avg_score": sum(scores) / len(scores),
            "ratings": ratings,
        }
